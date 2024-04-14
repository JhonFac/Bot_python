import csv
import io
import os

import google.generativeai as genai
import openpyxl
import PIL.Image
import PyPDF2
from docx import Document  # Para documentos Word
from dotenv import load_dotenv
from flask import Flask, jsonify, request
from flask_cors import CORS

load_dotenv()

# Configura el modelo y la API KEY
GOOGLE_API_KEY = 'AIzaSyCQ-pNzRJ0SaE-zcvDFJZZ2Oj3UX_ClyiA'
genai.configure(api_key=GOOGLE_API_KEY)

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}}, supports_credentials=True)  # Configuración CORS para permitir cualquier origen

@app.route('/analizar', methods=['POST'])
def analizar_imagen():
    # Verificar si se ha enviado algún archivo en la solicitud
    
    filtro = request.form.get('filtro', '')   
    
    if  filtro == 'MI MASCOTA':     
        
        if 'imagen' not in request.files:
            print(request.files)
            return jsonify({'mensaje': 'No se ha enviado ninguna imagen'}), 200
        
        imagen = request.files['imagen']
        
        # Comprobar si el archivo tiene un nombre y si es una imagen válida
        if imagen.filename == '' or not imagen.filename.endswith(('.jpg', '.jpeg', '.png')):
            return jsonify({'error': 'Archivo de imagen no válido'}), 400
        
        # Si se ha recibido una imagen, procesarla como lo hacías antes
        img = PIL.Image.open(imagen)
        modelo = genai.GenerativeModel('gemini-pro-vision')
        respuesta = modelo.generate_content(['Analiza la imagen y dime el estado de animo de la mascota, respuestas asi: Estado de Animo : ?, Raza o nombre: ? ,  Como se identifico el estado de animo:?,   Recomendaciones : ? , Biografia: ?  , Tipo (animal, humano, objetos, vehiculo): ? ,  las respuestas son en ESPAÑOL , entregame las respuestas siempre en un JSON y el JSON debe ir en orden como te entrege ,', img], stream=True)
        respuesta.resolve()  # Esperar a que la generación se complete
        respuesta = respuesta.text.replace('```json','').replace('```','')    
        return (respuesta)
    
    elif filtro == 'EJERCICIOS':   
          
        tipo = request.form.get('tipo', '')  
        
        if tipo == 'imagen':  
            if 'imagen' not in request.files:
                    print(request.files)
                    return jsonify({'mensaje': 'No se ha enviado ninguna imagen'}), 200
            
            imagen = request.files['imagen']
            
            # Comprobar si el archivo tiene un nombre y si es una imagen válida
            if imagen.filename == '' or not imagen.filename.endswith(('.jpg', '.jpeg', '.png')):
                return jsonify({'error': 'Archivo de imagen no válido'}), 400            
            # Si se ha recibido una imagen, procesarla como lo hacías antes
            try:
                img = PIL.Image.open(imagen)
                modelo = genai.GenerativeModel('gemini-pro-vision')
                respuesta = modelo.generate_content(['Analiza la imagen y resuelve el ejercicio de forma correcta, respuestas asi: Tipo de ejercicio: ?, Respuesta correcta:  ? , Paso a paso de la solución:?,      Recomendaciones : ? , Tipo (animal, humano, objetos, vehiculo, preguntas, ejercicios): ? , las respuestas son en ESPAÑOL , entregame las respuestas siempre en un JSON y el JSON debe ir en orden como te entrege', img], stream=True)
                respuesta.resolve()  # Esperar a que la generación se complete
                respuesta = respuesta.text.replace('```json','').replace('```','')    
                return (respuesta)
            except Exception as e:     
              return jsonify({'mensaje': 'Ups, lo siento. No se pudo procesar la solicitud. Inténtalo de nuevo.'}), 200
          
        elif tipo == 'texto':   
          try:            
            promtp = request.form.get('promtp', '')   
            modelo = genai.GenerativeModel('gemini-pro')
            respuesta = modelo.generate_content(f'Analiza este ejercicio ({promtp}) y resuelve el ejercicio de forma correcta, respuestas asi: Tipo de ejercicio: ?, Respuesta correcta:  ? , Paso a paso de la solución:?,      Recomendaciones : ? , Tipo : Ejercicio , las respuestas son en ESPAÑOL , entregame las respuestas siempre en un JSON y el JSON debe ir en orden como te entrege')
            respuesta.resolve()  # Esperar a que la generación se complete
            respuesta = respuesta.text.replace('```json','').replace('```','')    
            print(respuesta)
            return (respuesta)
          except Exception as e:     
              return jsonify({'mensaje': 'Ups, lo siento. No se pudo procesar la solicitud. Inténtalo de nuevo.'}), 200  
    
    elif filtro == 'AUTOMOVILES': 
        
        tipo = request.form.get('tipo', '')          
        if tipo == 'imagen':  
                    
            if 'imagen' not in request.files:
                    print(request.files)
                    return jsonify({'mensaje': 'No se ha enviado ninguna imagen'}), 200
            
            imagen = request.files['imagen']
            
            # Comprobar si el archivo tiene un nombre y si es una imagen válida
            if imagen.filename == '' or not imagen.filename.endswith(('.jpg', '.jpeg', '.png')):
                return jsonify({'error': 'Archivo de imagen no válido'}), 400
            
            # Si se ha recibido una imagen, procesarla como lo hacías antes
            try:
                img = PIL.Image.open(imagen)
                modelo = genai.GenerativeModel('gemini-pro-vision')
                respuesta = modelo.generate_content(['Analiza la imagen para que me respondas lo siguiente, respuestas asi: Marca del Vehiculo (Mazda 3, Chevrolet Camaro, Renault Twingo, etc): ?, , Modelo (Año) :? Precio del Vehiculo : ?  , Motor: ? , Potencia : ?  , Torque : ? , Transmision : ? , Rendimiento : ?, 0 a 100 en cuanto :?  , Gasolina que necesita en Octanos: ? , Biografia : ? , tipo (Carro, Moto, Bicicleta, Avion, etc): ? , las respuestas son en ESPAÑOL , entregame las respuestas siempre en un JSON y el JSON debe ir en orden como te entrege y no incluir JSON dentro del JSON solo un JSON como respuesta ', img], stream=True)
                respuesta.resolve()  # Esperar a que la generación se complete
                respuesta = respuesta.text.replace('```json','').replace('```','')    
                return (respuesta)
            except Exception as e:     
              return jsonify({'mensaje': 'Ups, lo siento. No se pudo procesar la solicitud. Inténtalo de nuevo.'}), 200
        
        elif tipo == 'texto':   
          try:            
            promtp = request.form.get('promtp', '')   
            modelo = genai.GenerativeModel('gemini-pro')
            respuesta = modelo.generate_content(f'Busca esto ({promtp})  para que me respondas lo siguiente, respuestas asi: Marca del Vehiculo (Mazda 3, Chevrolet Camaro, Renault Twingo, etc): ?, , Modelo (Año) :? Precio del Vehiculo : ?  , Motor: ? , Potencia : ?  , Torque : ? , Transmision : ? , Rendimiento : ?, 0 a 100 en cuanto :?  , Gasolina que necesita en Octanos: ? , Biografia : ? , tipo (Carro, Moto, Bicicleta, Avion, etc): ? , las respuestas son en ESPAÑOL , entregame las respuestas siempre en un JSON y el JSON debe ir en orden como te entrege y no incluir JSON dentro del JSON solo un JSON como respuesta ')
            respuesta.resolve()  # Esperar a que la generación se complete
            respuesta = respuesta.text.replace('```json','').replace('```','')    
            print(respuesta)
            return (respuesta)
          except Exception as e:     
              return jsonify({'mensaje': 'Ups, lo siento. No se pudo procesar la solicitud. Inténtalo de nuevo.'}), 200      
    
    
    elif filtro == 'IMAGEN A TEXTO':             
        if 'imagen' not in request.files:
                print(request.files)
                return jsonify({'mensaje': 'No se ha enviado ninguna imagen'}), 200
        
        imagen = request.files['imagen']
        
        # Comprobar si el archivo tiene un nombre y si es una imagen válida
        if imagen.filename == '' or not imagen.filename.endswith(('.jpg', '.jpeg', '.png')):
            return jsonify({'error': 'Archivo de imagen no válido'}), 400
        
        # Si se ha recibido una imagen, procesarla como lo hacías antes
        try:  
            img = PIL.Image.open(imagen)
            modelo = genai.GenerativeModel('gemini-pro-vision')
            respuesta = modelo.generate_content(['Extrae solo el texto de la Imagen , respuestas asi:   Extracción : Exitosa, Resultado (todo el  Texto extraido): ? , Explica el texto extraido: ? las respuestas son en ESPAÑOL , entregame las respuestas siempre en un JSON y el JSON debe ir en orden como te entrege y no incluir JSON dentro del JSON solo un JSON como respuesta ', img], stream=True)
            respuesta.resolve()  # Esperar a que la generación se complete
            respuesta = respuesta.text.replace('```json','').replace('```','')    
            return (respuesta)
        except Exception as e:     
                 return jsonify({'mensaje': 'Ups, lo siento. No se pudo procesar la solicitud. Inténtalo de nuevo.'}), 200

    elif filtro == 'FAMOSOS':  
        
        tipo = request.form.get('tipo', '')   
        
        if tipo == 'imagen':                   
            if 'imagen' not in request.files:
                    print(request.files)
                    return jsonify({'mensaje': 'No se ha enviado ninguna imagen'}), 200
            
            imagen = request.files['imagen']
            
            # Comprobar si el archivo tiene un nombre y si es una imagen válida
            if imagen.filename == '' or not imagen.filename.endswith(('.jpg', '.jpeg', '.png')):
                return jsonify({'error': 'Archivo de imagen no válido'}), 400
            
            # Si se ha recibido una imagen, procesarla como lo hacías antes
            try:  
                img = PIL.Image.open(imagen) 
                modelo = genai.GenerativeModel('gemini-pro-vision')
                respuesta = modelo.generate_content(['Analiza la imagen y encuentra al fomoso y sus redes sociales, respuestas asi: Nombre: ?,   Ocupacion (Influencer, Futbolista, Actor) :  ? , Instagram URl completa (Si no tiene Intagram poner NO TIENE) : ? ,  Facebook URl completa  (Si no tiene Facebook poner NO TIENE)  : ? ,  Tiktok URl completa  (Si no tiene Tiktok poner NO TIENE)  : ?  ,  Twitter URl completa  (Si no tiene Twitter poner NO TIENE)  : ? , Biografia completa  del famoso  : ? , las respuestas son en ESPAÑOL , entregame las respuestas siempre en un JSON y el JSON debe ir en orden como te entrege y no incluir JSON dentro del JSON solo un JSON como respuesta  ,', img], stream=True)
                respuesta.resolve()  # Esperar a que la generación se complete
                respuesta = respuesta.text.replace('```json','').replace('```','')    
                return (respuesta)
            except Exception as e:     
              return jsonify({'mensaje': 'Ups, lo siento. No se pudo procesar la solicitud. Inténtalo de nuevo.'}), 200
        
        
        elif tipo == 'texto':   
          try:            
            promtp = request.form.get('promtp', '')   
            modelo = genai.GenerativeModel('gemini-pro')
            respuesta = modelo.generate_content(f'Encuentra a {promtp} y sus redes sociales, respuestas asi: Nombre: ?,   Ocupacion (Influencer, Futbolista, Actor) :  ? , Instagram URl completa (Si no tiene Intagram poner NO TIENE) : ? ,  Facebook URl completa  (Si no tiene Facebook poner NO TIENE)  : ? ,  Tiktok URl completa  (Si no tiene Tiktok poner NO TIENE)  : ?  ,  Twitter URl completa  (Si no tiene Twitter poner NO TIENE)  : ? , Biografia completa  del famoso  : ? , las respuestas son en ESPAÑOL , entregame las respuestas siempre en un JSON y el JSON debe ir en orden como te entrege y no incluir JSON dentro del JSON solo un JSON como respuesta   ,', stream=True)
            respuesta.resolve()  # Esperar a que la generación se complete
            respuesta = respuesta.text.replace('```json','').replace('```','')    
            print(respuesta)
            return (respuesta)
          except Exception as e:     
              return jsonify({'mensaje': 'Ups, lo siento. No se pudo procesar la solicitud. Inténtalo de nuevo.'}), 200
            
            
    
    elif filtro == 'PELICULAS':    
        tipo = request.form.get('tipo', '')  
        if tipo == 'imagen':             
            if 'imagen' not in request.files:
                    print(request.files)
                    return jsonify({'mensaje': 'No se ha enviado ninguna imagen'}), 200
            
            imagen = request.files['imagen']
            
            # Comprobar si el archivo tiene un nombre y si es una imagen válida
            if imagen.filename == '' or not imagen.filename.endswith(('.jpg', '.jpeg', '.png')):
                return jsonify({'error': 'Archivo de imagen no válido'}), 400
            
            # Si se ha recibido una imagen, procesarla como lo hacías antes
            try: 
                img = PIL.Image.open(imagen) 
                modelo = genai.GenerativeModel('gemini-pro-vision')
                respuesta = modelo.generate_content(['Analiza la imagen y busca la pelicula, respuestas asi: Titulo: ?,   Lanzamiento  :  ? , Genero : ? ,  Actores principales : ?, Director : ? , ID de IMDB : URL completa con IMDB  , Plataforma  : URL Completa de la plataforma (Amazon, Neftlix, HBO, etc)   ,  Enlace Externo  : debe ser con esta url  https://streamsito.com/video/+ID de IMDB y conctaneas el ID de IMDB ? , Sipnosis  : ? , las respuestas son en ESPAÑOL , entregame las respuestas siempre en un JSON y el JSON debe ir en orden como te entrege y no incluir JSON dentro del JSON solo un JSON como respuesta ', img], stream=True)
                respuesta.resolve()  # Esperar a que la generación se complete
                respuesta = respuesta.text.replace('```json','').replace('```','')    
                return (respuesta)
            except Exception as e:     
              return jsonify({'mensaje': 'Ups, lo siento. No se pudo procesar la solicitud. Inténtalo de nuevo.'}), 200
        
        elif tipo == 'texto':   
          try:            
            promtp = request.form.get('promtp', '')   
            modelo = genai.GenerativeModel('gemini-pro')
            respuesta = modelo.generate_content(f'busca la pelicula {promtp}, y dame las respuestas asi: Titulo: ?,   Lanzamiento  :  ? , Genero : ? ,  Actores principales : ?, Director : ? , ID de IMDB : URL completa con IMDB  , Plataforma  : URL Completa de la plataforma (Amazon, Neftlix, HBO, etc)   ,  Enlace Externo  : debe ser con esta url  https://streamsito.com/video/ID de IMDB y conctaneas el ID de IMDB ? , Sipnosis  : ? , las respuestas son en ESPAÑOL , entregame las respuestas siempre en un JSON y el JSON debe ir en orden como te entrege y no incluir JSON dentro del JSON solo un JSON como respuesta  ')
            respuesta.resolve()  # Esperar a que la generación se complete
            respuesta = respuesta.text.replace('```json','').replace('```','')       
            return (respuesta)
          except Exception as e:     
              return jsonify({'mensaje': 'Ups, lo siento. No se pudo procesar la solicitud. Inténtalo de nuevo.'}), 200

    elif filtro == 'RESETAS':  
        tipo = request.form.get('tipo', '')   
        
        if tipo == 'imagen':             
            if 'imagen' not in request.files:
                    print(request.files)
                    return jsonify({'mensaje': 'No se ha enviado ninguna imagen'}), 200
            
            imagen = request.files['imagen']
            modelo = genai.GenerativeModel('gemini-pro-vision')
            # Comprobar si el archivo tiene un nombre y si es una imagen válida
            if imagen.filename == '' or not imagen.filename.endswith(('.jpg', '.jpeg', '.png')):
                return jsonify({'error': 'Archivo de imagen no válido'}), 400
            
            try: 
                # Si se ha recibido una imagen, procesarla como lo hacías antes
                img = PIL.Image.open(imagen) 
                
                respuesta = modelo.generate_content(['Analiza la imagen la cual tiene ingedientes para preparar una reseta, respuestas asi: Titulo reseta: ?, Lista de ingredientes :  ? , Preparación :  ? ,  Tiempo de cocción : ?, Observaciones : ? , las respuestas son en ESPAÑOL , entregame las respuestas siempre en un JSON y el JSON debe ir en orden como te entrege y no incluir JSON dentro del JSON solo un JSON como respuesta ', img], stream=True)
                respuesta.resolve()  # Esperar a que la generación se complete
                respuesta = respuesta.text.replace('```json','').replace('```','')    
                return (respuesta)
            except Exception as e:     
              return jsonify({'mensaje': 'Ups, lo siento. No se pudo procesar la solicitud. Inténtalo de nuevo.'}), 200
        
        elif tipo == 'texto':   
          try:            
            promtp = request.form.get('promtp', '')   
            modelo = genai.GenerativeModel('gemini-pro')
            respuesta = modelo.generate_content(f'Analiza esto {promtp} y dame una reseta de cocina, respuestas asi: Titulo reseta: ?, Lista de ingredientes :  ? , Preparación :  ? ,  Tiempo de cocción : ?, Observaciones : ? , las respuestas son en ESPAÑOL , entregame las respuestas siempre en un JSON y el JSON debe ir en orden como te entrege y no incluir JSON dentro del JSON solo un JSON como respuesta ')
            respuesta.resolve()  # Esperar a que la generación se complete
            respuesta = respuesta.text.replace('```json','').replace('```','')       
            return (respuesta)
          except Exception as e:     
              return jsonify({'mensaje': 'Ups, lo siento. No se pudo procesar la solicitud. Inténtalo de nuevo.'}), 200
    
    elif filtro == 'DOCUMENTOS':             
        if 'archivo' not in request.files:
            print(request.files)
            return jsonify({'mensaje': 'No se ha enviado ningún archivo'}), 200
    
        archivo = request.files['archivo']
        promtp = request.form.get('promtp', '')   
        nombre_archivo = archivo.filename
               
        if nombre_archivo.lower().endswith('.pdf'):           
            texto = extraer_texto_pdf(archivo.stream)
    # Procesa el texto extraído del PDF
        elif nombre_archivo.lower().endswith(('.docx', '.doc' )):           
            texto = extraer_texto_docx(archivo.stream)
            # Procesa el texto extraído del documento Word
        elif nombre_archivo.lower().endswith('.txt'):
              texto = extraer_texto_txt(archivo)              
        elif nombre_archivo.lower().endswith('.csv'):
             texto = extraer_texto_csv(archivo)   
        elif nombre_archivo.lower().endswith('.xlsx'):
             texto = extraer_texto_xlsx(archivo)                
        else:
            return jsonify({'mensaje': 'El archivo enviado no es un formato soportado'}), 400
        
        if texto == "El documento está protegido con contraseña.":
                 return ("El documento está protegido con contraseña, no es posible acceder al documento")
        else:            
           
            try:
                modelo = genai.GenerativeModel('gemini-pro')
                respuesta = modelo.generate_content(f'Analiza lo siguiente ({texto}), ahora responde lo siguiente:{promtp}, responde muy natural como si fuera una persona quien haya leído el texto, evita el plagio, las respuestas son en ESPAÑOL.')
                respuesta.resolve()  # Esperar a que la generación se complete    
                respuesta = respuesta.text.replace('```json','').replace('```','')            
                return (respuesta)
            except Exception as e:
                
                return "Ups, lo siento, hay un error con el archivo. No se pudo procesar la solicitud. Inténtalo de nuevo."

         
      
    
def extraer_texto_docx(archivo):
    doc = Document(archivo)
    texto_completo = []
    for para in doc.paragraphs:
        texto_completo.append(para.text)
    archivo.seek(0)  # Restablece el puntero del archivo
    return '\n'.join(texto_completo)  

def extraer_texto_pdf(archivo):
    lector_pdf = PyPDF2.PdfReader(archivo)
    texto = ""
    
    # Verificar si el PDF está protegido con contraseña
    if lector_pdf.is_encrypted:
        texto = 'El documento está protegido con contraseña.'      
    else:
        for pagina in range(len(lector_pdf.pages)):
            texto += lector_pdf.pages[pagina].extract_text()

    archivo.seek(0)  # Restablece el puntero del archivo
    return texto
    
def extraer_texto_txt(archivo):
    texto_completo = []
    for linea in archivo:
        texto_completo.append(linea.decode('utf-8').strip())
    archivo.seek(0)  # Restablece el puntero del archivo
    return '\n'.join(texto_completo)

def extraer_texto_csv(archivo):
    texto_completo = []
    
    # Crear un objeto TextIOWrapper para leer el archivo CSV en modo texto
    archivo_texto = io.TextIOWrapper(archivo, encoding='utf-8', newline='')
    
    csv_reader = csv.reader(archivo_texto)
    for fila in csv_reader:
        texto_completo.extend(fila)
    
    archivo.seek(0)  # Restablece el puntero del archivo
    return '\n'.join(texto_completo)

def extraer_texto_xlsx(archivo):
    texto_completo = []
    workbook = openpyxl.load_workbook(archivo)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    texto_completo.append(str(cell.value))
    return '\n'.join(texto_completo)   
    
if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)